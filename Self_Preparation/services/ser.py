import psutil
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Define process names (adjust as needed)
process_names = {
    "MIMSMonitoringService": "MIMSMonitoringService",
    "AbilityMACHConnectService": "AbilityMACHConnectService",
    "AbilityMACHEventConnectService": "AbilityMACHEventConnectService",  # Replace with actual DB process name
    "Dnscache": "Dnscache",
    "Dhcp":"Dhcp"
}

excel_file = "detailed_service_metrics.xlsx"


def get_process_metrics(proc_name):
    cpu = 0.0
    mem = 0.0
    for proc in psutil.process_iter(['name', 'cpu_percent', 'memory_info']):
        try:
            if proc.info['name'].lower() == proc_name.lower():
                cpu += proc.cpu_percent(interval=0.1)
                mem += proc.info['memory_info'].rss / (1024 * 1024)  # MB
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue
    return cpu, mem


def write_to_excel(timestamp, metrics):
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Service Metrics"
        headers = ["Timestamp"]
        for service in process_names:
            headers.extend([f"{service} CPU (%)", f"{service} Memory (MB)"])
        ws.append(headers)
    else:
        wb = load_workbook(excel_file)
        ws = wb["Service Metrics"]

    row = [timestamp]
    for service in process_names:
        cpu, mem = metrics.get(service, (0.0, 0.0))
        row.extend([round(cpu, 2), round(mem, 2)])

    ws.append(row)
    wb.save(excel_file)


if __name__ == "__main__":
    while True:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        metrics = {}
        for label, proc_name in process_names.items():
            cpu, mem = get_process_metrics(proc_name)
            metrics[label] = (cpu, mem)
        write_to_excel(timestamp, metrics)
        time.sleep(2)