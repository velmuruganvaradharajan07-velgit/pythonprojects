import win32serviceutil
import win32service
import psutil
import wmi
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import os
""" Services need to be updated to get the cpu, memory and status of the service"""
services_to_monitor = [
    "MIMSMonitoringService",
    "AbilityMACHConnectService",
    "AbilityMACHEventConnectService",
    "Spooler",
    "wuauserv"
]

c = wmi.WMI()
excel_file = "services_monitoring.xlsx"

"""Creation of excel sheet """
if not os.path.exists(excel_file):
    wb = Workbook()
    for i, service_name in enumerate(services_to_monitor):
        ws = wb.active if i == 0 else wb.create_sheet(title=service_name)
        ws.title = service_name
        ws.append(["Timestamp", "Status", "PID", "CPU Usage (%)", "Memory Usage (MB)"])
    wb.save(excel_file)

while True:
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(excel_file)

    for service_name in services_to_monitor:
        try:
            status = win32serviceutil.QueryServiceStatus(service_name)[1]
            status_map = {
                win32service.SERVICE_STOPPED: "Stopped",
                win32service.SERVICE_START_PENDING: "Start Pending",
                win32service.SERVICE_STOP_PENDING: "Stop Pending",
                win32service.SERVICE_RUNNING: "Running",
                win32service.SERVICE_CONTINUE_PENDING: "Continue Pending",
                win32service.SERVICE_PAUSE_PENDING: "Pause Pending",
                win32service.SERVICE_PAUSED: "Paused"
            }
            status_text = status_map.get(status, "Unknown")
        except Exception as e:
            status_text = f"Error: {e}"

        pid = None
        for service in c.Win32_Service(Name=service_name):
            pid = service.ProcessId

        cpu_usage = "N/A"
        mem_usage_mb = "N/A"

        if pid and pid != 0:
            try:
                proc = psutil.Process(pid)
                cpu_usage = proc.cpu_percent(interval=1)
                mem_info = proc.memory_info()
                mem_usage_mb = mem_info.rss / (1024 * 1024)
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass

        ws = wb[service_name]
        ws.append([timestamp, status_text, pid, cpu_usage, mem_usage_mb])

        # Plotting
        timestamps = [row[0].value for row in ws.iter_rows(min_row=2, max_col=1)]
        cpu_values = [row[3].value if isinstance(row[3].value, (int, float)) else 0 for row in ws.iter_rows(min_row=2, max_col=4)]
        mem_values = [row[4].value if isinstance(row[4].value, (int, float)) else 0 for row in ws.iter_rows(min_row=2, max_col=5)]

        plt.figure(figsize=(10, 5))
        plt.plot(timestamps, cpu_values, marker='o')
        plt.title(f"CPU Usage - {service_name}")
        plt.xlabel("Timestamp")
        plt.ylabel("CPU Usage (%)")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(f"{service_name}_cpu_usage.png")
        plt.close()

        plt.figure(figsize=(10, 5))
        plt.plot(timestamps, mem_values, marker='o', color='green')
        plt.title(f"Memory Usage - {service_name}")
        plt.xlabel("Timestamp")
        plt.ylabel("Memory Usage (MB)")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(f"{service_name}_memory_usage.png")
        plt.close()

    wb.save(excel_file)
    time.sleep(5)  # Wait 60 seconds before next update